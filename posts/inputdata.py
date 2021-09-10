import json
import os

import pymysql
import openpyxl
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView

from bases.response import APIResponse
from posts.models import Region

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

with open(os.path.join(BASE_DIR, 'secrets.json'), 'rb') as secret_file:
    secrets = json.load(secret_file)


# wb = openpyxl.load_workbook(filename="C:/Users/최예원/Downloads/region.xlsx")
# ws = wb.active

# input 완료
class InputRegionView(APIView):
    permission_classes = [AllowAny, ]

    def post(self, request):
        response = APIResponse(success=False, code=400)
        host = secrets["DATABASES"]["default"]["HOST"]
        user = secrets["DATABASES"]["default"]["USER"]
        pw = secrets["DATABASES"]["default"]["PASSWORD"]

        conn = pymysql.connect(host=host, user=user, password=pw, db='carping')
        curs = conn.cursor(pymysql.cursors.DictCursor)

        for i in range(2, 18497):
            for j in range(1, 4):
                id = i - 1
                # ws.cell(row=i, column=j)
                sido = ws.cell(row=i, column=1).value
                sigungu = ws.cell(row=i, column=2).value
                dong = ws.cell(row=i, column=3).value

                if Region.objects.filter(sido=sido, sigungu=sigungu, dong=dong):
                    pass
                else:
                    sql = """INSERT INTO posts_region(id, sido, sigungu, dong) VALUES(%s, %s, %s, %s)"""
                    curs.execute(sql, (id, sido, sigungu, dong))
                    conn.commit()
        response.success = True
        response.code = 200
        return response.response(data=[{"message": "completed"}])
